import os
import re
import csv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# Define paths
schemas_dir = 'individual_schemas'
data_dir = 'individual_schemas_with_data'

# Get all sql files
schema_files = [f for f in os.listdir(schemas_dir) if f.endswith('.sql') and f != '00_enums.sql']
data_files = [f for f in os.listdir(data_dir) if f.endswith('.sql')]

tables = set([f[:-4] for f in schema_files])
tables.update([f[:-4] for f in data_files])

print(f"Total unique tables found: {len(tables)}")

def parse_sql_file(filepath):
    if not os.path.exists(filepath):
        return None, []
    with open(filepath, 'r') as f:
        content = f.read()

    table_match = re.search(r'CREATE TABLE\s+(\w+)\s*\((.*?)\);', content, re.DOTALL | re.IGNORECASE)

    if not table_match:
        return None, []

    table_name = table_match.group(1)
    columns_str = table_match.group(2)

    parts = re.split(r',\s*(?![^()]*\))', columns_str.strip())

    columns = []
    for part in parts:
        part = part.strip()
        if not part or part.upper().startswith(('PRIMARY KEY', 'FOREIGN KEY', 'UNIQUE', 'CHECK', 'CONSTRAINT')):
            continue

        col_parts = part.split(maxsplit=2)
        if len(col_parts) >= 2:
            col_name = col_parts[0]
            col_type = col_parts[1]
            constraints = col_parts[2] if len(col_parts) > 2 else ""
            columns.append({
                'name': col_name,
                'type': col_type,
                'constraints': constraints
            })

    return table_name, columns

def parse_values_string(s):
    rows = []
    current_row = []
    current_val = []
    in_string = False
    paren_depth = 0
    i = 0
    while i < len(s):
        c = s[i]
        if c == "'":
            if in_string and i + 1 < len(s) and s[i+1] == "'":
                current_val.append("''")
                i += 2
                continue
            else:
                in_string = not in_string
                current_val.append(c)
        elif in_string:
            current_val.append(c)
        else:
            if c == '(':
                paren_depth += 1
                if paren_depth > 1:
                    current_val.append(c)
            elif c == ')':
                paren_depth -= 1
                if paren_depth > 0:
                    current_val.append(c)
                elif paren_depth == 0:
                    current_row.append("".join(current_val).strip())
                    current_val = []
                    rows.append(current_row)
                    current_row = []
            elif c == ',' and paren_depth == 1:
                current_row.append("".join(current_val).strip())
                current_val = []
            elif paren_depth > 0:
                current_val.append(c)
        i += 1
    return rows

def parse_insert_data(filepath):
    if not os.path.exists(filepath):
        return []
    with open(filepath, 'r') as f:
        content = f.read()

    inserts = []
    statements = re.split(r'INSERT INTO\s+', content, flags=re.IGNORECASE)[1:]
    for stmt in statements:
        match = re.match(r'(\w+)\s*\((.*?)\)\s*VALUES\s*(.*?);', stmt, re.DOTALL | re.IGNORECASE)
        if match:
            table_name = match.group(1)
            columns = [c.strip() for c in match.group(2).split(',')]
            values_str = match.group(3).strip()
        else:
            match2 = re.match(r'(\w+)\s*VALUES\s*(.*?);', stmt, re.DOTALL | re.IGNORECASE)
            if match2:
                table_name = match2.group(1)
                columns = []
                values_str = match2.group(2).strip()
            else:
                continue

        rows = parse_values_string(values_str)

        inserts.append({
            'table': table_name,
            'columns': columns,
            'rows': rows
        })
    return inserts

wb = Workbook()
wb.remove(wb.active)

def safe_sheet_name(name):
    return name[:31]

for table in sorted(tables):
    sheet_name = safe_sheet_name(table)
    idx = 1
    original_sheet_name = sheet_name
    while sheet_name in wb.sheetnames:
        sheet_name = f"{original_sheet_name[:28]}_{idx}"
        idx += 1
    ws = wb.create_sheet(title=sheet_name)

    schema_path = os.path.join(schemas_dir, f"{table}.sql")
    data_path = os.path.join(data_dir, f"{table}.sql")

    t_name, cols = parse_sql_file(schema_path)
    if not cols:
        t_name, cols = parse_sql_file(data_path)

    if cols:
        ws.append([col['name'] for col in cols])
        ws.append([col['type'] for col in cols])
        ws.append([col['constraints'] for col in cols])

        inserts = parse_insert_data(data_path)
        inserts.extend(parse_insert_data(schema_path))

        if inserts:
            col_names = [col['name'] for col in cols]
            for insert in inserts:
                insert_cols = insert['columns']
                for row in insert['rows']:
                    out_row = []
                    for c_name in col_names:
                        if insert_cols and c_name in insert_cols:
                            c_idx = insert_cols.index(c_name)
                            if c_idx < len(row):
                                out_row.append(row[c_idx])
                            else:
                                out_row.append("")
                        elif not insert_cols:
                            c_idx = col_names.index(c_name)
                            if c_idx < len(row):
                                out_row.append(row[c_idx])
                            else:
                                out_row.append("")
                        else:
                            out_row.append("")
                    ws.append(out_row)
    else:
        ws.append([f"No schema found for {table}"])

wb.save("Spreadsheet_editable.xlsx")
print("Done!")
